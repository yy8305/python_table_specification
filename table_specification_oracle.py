#################################################
# oracle 테이블 명세서 엑셀로 생성
#
#################################################


from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import cx_Oracle
import os
os.putenv("NLS_LANG", "KOREAN_KOREA.KO16KSC5601")#한글깨짐

write_wb = Workbook()

# 이름이 있는 시트를 생성
write_ws = write_wb.create_sheet('Sheet1')

# Sheet1에다 입력
write_ws = write_wb.active



grayFill = PatternFill(start_color='c0c0c0',end_color='c0c0c0',fill_type='solid')
box = Border(left=Side(border_style="thin",
                   color='FF000000'),
         right=Side(border_style="thin",
                    color='FF000000'),
         top=Side(border_style="thin",
                  color='FF000000'),
         bottom=Side(border_style="thin",
                     color='FF000000'),
         diagonal=Side(border_style="thin",
                       color='FF000000'),
         diagonal_direction=0,
         outline=Side(border_style="thin",
                      color='FF000000'),
         vertical=Side(border_style="thin",
                       color='FF000000'),
         horizontal=Side(border_style="thin",
                        color='FF000000')
        )



# Oracle 서버와 연결(Connection 맺기)
conn = cx_Oracle.connect('[User_name]/[Password]@[Server IP]:[Port]/[Service Name]')
cursor = conn.cursor() # cursor 객체 얻어오기
cursor.execute("""
SELECT 
  A.OWNER, A.table_name, B.comments, A.NUM_ROWS
FROM 
  all_tables A, all_tab_comments B
WHERE 
  A.OWNER = '[User]' 
  AND A.table_name = B.table_name
  AND (A.NUM_ROWS > 0 or A.NUM_ROWS is null)
GROUP BY 
  A.OWNER, A.table_name, B.comments, A.NUM_ROWS
ORDER by 
  A.TABLE_NAME
""") # SQL 문장 실행
rows = cursor.fetchall()


i=1
for row in rows:
    ############################1행##########################
    write_ws.merge_cells("A"+str(i)+":I"+str(i)); # 병합
    write_ws["A"+str(i)+""] = '테이블명세서'
    write_ws["A"+str(i)+""].fill = grayFill # 배경색
    write_ws["A"+str(i)+""].alignment = Alignment(horizontal='center', vertical='center')
    write_ws["A"+str(i)+""].font = Font(name='굴림체', size=11, bold=True)
    for x in range(1,10): # border
        write_ws.cell(row=i, column=x).border = box

    ############################2행##########################
    write_ws.merge_cells("A"+str(i+1)+":B"+str(i+1)); # 병합
    write_ws["A"+str(i+1)+""] = '작성일'
    write_ws["A"+str(i+1)+""].fill = grayFill # 배경색
    write_ws["A"+str(i+1)+""].alignment = Alignment(horizontal='center', vertical='center')
    write_ws["A"+str(i+1)+""].font = Font(name='굴림체', size=11, bold=True)

    write_ws["C"+str(i+1)+""] = '2019-09-09'

    write_ws.merge_cells("D"+str(i+1)+":E"+str(i+1)); # 병합
    write_ws["D"+str(i+1)+""] = '작성자'
    write_ws["D"+str(i+1)+""].fill = grayFill # 배경색
    write_ws["D"+str(i+1)+""].alignment = Alignment(horizontal='center', vertical='center')
    write_ws["D"+str(i+1)+""].font = Font(name='굴림체', size=11, bold=True)

    write_ws.merge_cells("F"+str(i+1)+":I"+str(i+1)); # 병합
    write_ws["F"+str(i+1)+""] = '박홍철'
    write_ws["F"+str(i+1)+""].alignment = Alignment(horizontal='center', vertical='center')

    for x in range(1,10): # border
        write_ws.cell(row=(i+1), column=x).border = box

    ############################3행##########################
    write_ws.merge_cells("A" + str(i + 2) + ":B" + str(i + 2));  # 병합
    write_ws["A" + str(i + 2) + ""] = '테이블ID'
    write_ws["A" + str(i + 2) + ""].fill = grayFill  # 배경색
    write_ws["A" + str(i + 2) + ""].alignment = Alignment(horizontal='center', vertical='center')
    write_ws["A" + str(i + 2) + ""].font = Font(name='굴림체', size=11, bold=True)

    write_ws["C" + str(i + 2) + ""] = row[1]

    write_ws.merge_cells("D" + str(i + 2) + ":E" + str(i + 2));  # 병합
    write_ws["D" + str(i + 2) + ""] = '테이블명'
    write_ws["D" + str(i + 2) + ""].fill = grayFill  # 배경색
    write_ws["D" + str(i + 2) + ""].alignment = Alignment(horizontal='center', vertical='center')
    write_ws["D" + str(i + 2) + ""].font = Font(name='굴림체', size=11, bold=True)

    write_ws.merge_cells("F" + str(i + 2) + ":I" + str(i + 2));  # 병합
    write_ws["F" + str(i + 2) + ""] = row[2]
    write_ws["F" + str(i + 2) + ""].alignment = Alignment(horizontal='center', vertical='center')

    for x in range(1, 10):  # border
        write_ws.cell(row=(i + 2), column=x).border = box


    ############################4행##########################
    write_ws.merge_cells("A" + str(i + 3) + ":B" + str(i + 3));  # 병합
    write_ws["A" + str(i + 3) + ""] = '테이블설명'
    write_ws["A" + str(i + 3) + ""].fill = grayFill  # 배경색
    write_ws["A" + str(i + 3) + ""].alignment = Alignment(horizontal='center', vertical='center')
    write_ws["A" + str(i + 3) + ""].font = Font(name='굴림체', size=11, bold=True)

    write_ws.merge_cells("C" + str(i + 3) + ":I" + str(i + 3));  # 병합

    for x in range(1, 10):  # border
        write_ws.cell(row=(i + 3), column=x).border = box


    ############################5행##########################
    write_ws["A" + str(i + 4) + ""] = 'No.'
    write_ws["B" + str(i + 4) + ""] = '컬럼ID'
    write_ws["C" + str(i + 4) + ""] = '컬럼명'
    write_ws["D" + str(i + 4) + ""] = '타입'
    write_ws["E" + str(i + 4) + ""] = '길이'
    write_ws["F" + str(i + 4) + ""] = 'NULL'
    write_ws["G" + str(i + 4) + ""] = 'KEY'
    write_ws["H" + str(i + 4) + ""] = 'DEFAULT'
    write_ws["I" + str(i + 4) + ""] = '비고'

    for x in range(1, 10):  # border
        write_ws.cell(row=(i + 4), column=x).border = box
        write_ws.cell(row=(i + 4), column=x).fill = grayFill  # 배경색
        write_ws.cell(row=(i + 4), column=x).alignment = Alignment(horizontal='center', vertical='center')
        write_ws.cell(row=(i + 4), column=x).font = Font(name='굴림체', size=11, bold=True)

    i = i+5
    ############################컬럼##########################
    conn2 = cx_Oracle.connect('[User_name]/[Password]@[Server IP]:[Port]/[Service Name]')
    cursor2 = conn2.cursor()  # cursor 객체 얻어오기
    cursor2.execute("""
        SELECT A1.TABLE_COMMENTS TABLE_COMMENTS
         , A1.TABLE_NAME TABLE_NAME
         , A1.COLUMN_COMMENTS COLUMN_COMMENTS
         , A1.COLUMN_NAME COLUMN_NAME
         , (CASE
               WHEN B1.CONSTRAINT_TYPE = 'P'
                  THEN 'Y'
            END) PK_FLAG
         , (CASE
               WHEN B1.CONSTRAINT_TYPE = 'R'
                  THEN 'Y'
            END) FK_FLAG
         , A1.NULL_FLAG
         , A1.DATA_TYPE
         , A1.DATA_LENGTH
      FROM (SELECT B.COMMENTS TABLE_COMMENTS
                 , A.TABLE_NAME TABLE_NAME
                 , C.COMMENTS COLUMN_COMMENTS
                 , A.COLUMN_NAME COLUMN_NAME
                 , (CASE A.NULLABLE
                       WHEN 'Y'
                          THEN 'Y'
                    END) NULL_FLAG
                 , A.DATA_TYPE DATA_TYPE
                 , (CASE
                       WHEN A.DATA_TYPE IN ('CHAR', 'VARCHAR2')
                          THEN '' || A.DATA_LENGTH || ''
                       WHEN A.DATA_TYPE = 'NUMBER'
                       AND A.DATA_SCALE = 0
                       AND A.DATA_PRECISION IS NOT NULL
                          THEN '' || A.DATA_PRECISION || ''
                       WHEN A.DATA_TYPE = 'NUMBER'
                       AND A.DATA_SCALE <> 0
                          THEN '' || A.DATA_PRECISION || ',' || A.DATA_SCALE
                               || ''
                    END
                   ) DATA_LENGTH
     
                 , A.COLUMN_ID
              FROM USER_TAB_COLUMNS A
                 , USER_TAB_COMMENTS B
                 , USER_COL_COMMENTS C
             WHERE (A.TABLE_NAME = B.TABLE_NAME)
               AND (    A.TABLE_NAME = C.TABLE_NAME
                    AND A.COLUMN_NAME = C.COLUMN_NAME
                   )) A1
         , (SELECT A.TABLE_NAME
                 , A.COLUMN_NAME
                 , B.CONSTRAINT_TYPE
              FROM USER_CONS_COLUMNS A
                 , USER_CONSTRAINTS B
             WHERE (A.CONSTRAINT_NAME = B.CONSTRAINT_NAME)
               AND B.CONSTRAINT_TYPE IN ('P', 'R')) B1
    WHERE (    A1.TABLE_NAME = B1.TABLE_NAME(+)
            AND A1.COLUMN_NAME = B1.COLUMN_NAME(+))
            AND A1.TABLE_NAME = '"""+row[1]+"""'
    ORDER BY A1.TABLE_NAME, A1.COLUMN_ID
    """)
    rows2 = cursor2.fetchall()
    j=1
    index = []
    for row2 in rows2:
        ret_val = -1
        if (len(index) > 0):
            for ind in index:
                try:
                    ret_val = ind[3].index(row2[3])
                except:
                    ret_val = -1
        if (ret_val < 0):
            write_ws["A" + str(i) + ""] = str(j)
            write_ws["B" + str(i) + ""] = row2[3]
            write_ws["C" + str(i) + ""] = row2[2]
            write_ws["D" + str(i) + ""] = row2[7]
            write_ws["E" + str(i) + ""] = row2[8]
            if(row2[6] == "Y"):
                write_ws["F" + str(i) + ""] = "NULL"
            else:
                write_ws["F" + str(i) + ""] = "Not NULL"

            if(row2[4] == "Y" and row2[5] != "Y"):
                write_ws["G" + str(i) + ""] = 'PK'
                index.append(row2)
            elif(row2[4] != "Y" and row2[5] == "Y"):
                write_ws["G" + str(i) + ""] = 'FK'
                index.append(row2)
            elif (row2[4] == "Y" and row2[5] == "Y"):
                write_ws["G" + str(i) + ""] = 'PK, FK'
                index.append(row2)

            for x in range(1, 10):  # border
                write_ws.cell(row=(i), column=x).border = box
                write_ws.cell(row=(i), column=x).alignment = Alignment(horizontal='center', vertical='center')

            i = i + 1
            j = j + 1
        else:
            write_ws["G" + str(i-1) + ""] = 'PK, FK'
            index.append(row2)

    ############################인덱스##########################
    write_ws.merge_cells("A" + str(i) + ":B" + str(i));  # 병합
    write_ws["A" + str(i) + ""] = '인덱스'

    write_ws.merge_cells("C" + str(i) + ":I" + str(i));  # 병합
    write_ws["C" + str(i) + ""] = '인덱스키'

    for x in range(1, 10):  # border
        write_ws.cell(row=(i), column=x).border = box
        write_ws.cell(row=(i), column=x).fill = grayFill  # 배경색
        write_ws.cell(row=(i), column=x).alignment = Alignment(horizontal='center', vertical='center')
        write_ws.cell(row=(i), column=x).font = Font(name='굴림체', size=11, bold=True)

    i = i+1

    if(len(index) <= 0):
        write_ws.merge_cells("A" + str(i) + ":B" + str(i));  # 병합
        write_ws.merge_cells("C" + str(i) + ":I" + str(i));  # 병합
        for x in range(1, 10):  # border
            write_ws.cell(row=(i), column=x).border = box
            write_ws.cell(row=(i), column=x).alignment = Alignment(horizontal='center', vertical='center')
        i = i+1

    index_pk = []
    index_fk = []
    for y in index:
        if(y[4] == "Y" and y[5] != "Y"):
            index_pk.append(y[3])
        elif (y[4] != "Y" and y[5] == "Y"):
            index_fk.append(y[3])
        elif (y[4] == "Y" and y[5] == "Y"):
            index_pk.append(y[3])
            index_fk.append(y[3])

    if (len(index_pk) > 0):
        write_ws.merge_cells("A" + str(i) + ":B" + str(i));  # 병합
        write_ws["A" + str(i) + ""] = str(row[1]) + "_PK"
        pk_str = ""
        for pk in index_pk:
            pk_str = pk_str + ", " + pk

        write_ws.merge_cells("C" + str(i) + ":I" + str(i));  # 병합
        write_ws["C" + str(i) + ""] = pk_str[1:]

        for x in range(1, 10):  # border
            write_ws.cell(row=(i), column=x).border = box
            write_ws.cell(row=(i), column=x).alignment = Alignment(horizontal='center', vertical='center')
        i = i + 1

    if (len(index_fk) > 0):
        write_ws.merge_cells("A" + str(i) + ":B" + str(i));  # 병합
        write_ws["A" + str(i) + ""] = str(row[1]) + "_fK"
        fk_str = ""
        for fk in index_fk:
            fk_str = fk_str + ", " + fk

        write_ws.merge_cells("C" + str(i) + ":I" + str(i));  # 병합
        write_ws["C" + str(i) + ""] = fk_str[1:]

        for x in range(1, 10):  # border
            write_ws.cell(row=(i), column=x).border = box
            write_ws.cell(row=(i), column=x).alignment = Alignment(horizontal='center', vertical='center')
        i = i + 1

    ############################업무규칙##########################
    write_ws["A" + str(i) + ""] = '업무규칙'
    write_ws["A" + str(i) + ""].border = box
    write_ws["A" + str(i) + ""].fill = grayFill  # 배경색
    write_ws["A" + str(i) + ""].alignment = Alignment(horizontal='center', vertical='center')
    write_ws["A" + str(i) + ""].font = Font(name='굴림체', size=11, bold=True)

    write_ws.merge_cells("B" + str(i) + ":I" + str(i));  # 병합

    for x in range(2, 10):  # border
        write_ws.cell(row=(i), column=x).border = box
        write_ws.cell(row=(i), column=x).alignment = Alignment(horizontal='center', vertical='center')

    i = i+1


    i = i+5

write_wb.save('C:/oracle_table_excel.xlsx') #해당 파일명으로 저장